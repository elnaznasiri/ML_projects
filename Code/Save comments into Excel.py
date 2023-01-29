import requests
from openpyxl import Workbook
from typing import List, TypedDict

# Type definitions
class Reactions(TypedDict):
    likes: int
    dislikes: int


class Comment(TypedDict):
    id: int
    title: str
    body: str
    created_at: str
    rate: int
    reactions: Reactions
    is_buyer: bool
    user_name: str
    is_anonymous: bool


def save_comments_to_exel(json_data, filename):
    keys = []
    wb = Workbook()
    ws = wb.active

    for i in range(len(json_data)):
        sub_obj = json_data[i]
        if i == 0:
            keys = list(sub_obj.keys())
            for k in range(len(keys)):
                ws.cell(row=(i + 1), column=(k + 1), value=keys[k])
        for j in range(len(keys)):
            try:
                ws.cell(row=(i + 2), column=(j + 1), value=sub_obj[keys[j]])
            except:
                pass

    wb.save(filename)


def get_comment(sess: requests.Session, product_id: int, page: int = 1, pager: bool = False) -> List[Comment]:
    URL = f'https://api.digikala.com/v1/product/{product_id}/comments/?page={page}'
    response = sess.get(URL)

    if response.ok:
        data = response.json().get('data')
        complete_comment = data.get('comments')

        comments = []
        for comment in complete_comment:
            comments.append({
                            'productID' : product_id,
                            'id': comment.get('id'),
                            'title': comment.get('title'),
                            'body': comment.get('body'),
                            'created_at': comment.get('created_at'),
                            'rate': comment.get('rate'),
                            'likes': comment.get('reactions').get('likes'),
                            'dislikes':comment.get('reactions').get('dislikes'),
                            'recommendation_status': comment.get('recommendation_status'),
                            'is_buyer': comment.get('is_buyer'),
                            'user_name': comment.get('user_name'),
                            'is_anonymous': comment.get('is_anonymous'),
                            })

        if pager:
            return  comments, data.get('pager')
        return comments
    return []


if __name__ == '__main__':
    product_id = 3493882
    
    sesstion = requests.Session()
    filename = "D:\Programming\ML_projects\OutputFiles\Stage\\" +  str(product_id)  + ".xlsx"
    comments, pager = get_comment(sesstion, product_id, pager=True)
    total_pages = pager.get('total_pages')

    for page in range(1, total_pages + 1):
        comment = get_comment(sesstion, product_id, page=page)
        comments.extend(comment)
    save_comments_to_exel(comments, filename)
